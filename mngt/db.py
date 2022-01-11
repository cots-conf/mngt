"""Database stuff.

:author: Krerkkiat Chusap
"""
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Tuple

import arrow
import click
from flask import Flask, current_app, g
from flask.cli import with_appcontext
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session

_users = {
    "kc555014@ohio.edu": {"password": "secret"},
}

_conferences = [
    {
        "name": "COTS 2021",
        "description": "2021 Meeting of the Council on Thai Studies (COTS)",
        "begin": arrow.get(datetime(2021, 11, 12, 16, 30), "US/Eastern")
        .to("utc")
        .datetime,
        "end": arrow.get(datetime(2021, 11, 13, 23, 45), "US/Eastern")
        .to("utc")
        .datetime,
    }
]

_participants = [
    {
        "email": "bobby@example.com",
        "first_name": "Bobby",
        "last_name": "B",
        "affiliation": "Example Inc.",
        "proposals": [
            {"title": "Lorem ipsum", "type": "paper", "abstract": "Lorem ipsum"},
            {"title": "Lorem ipsum", "type": "paper", "abstract": "Lorem ipsum"},
        ],
    },
    {
        "email": "water@example.com",
        "first_name": "Water",
        "last_name": "C",
        "affiliation": "MetaCorp Inc.",
        "proposals": [],
    },
    {
        "email": "janrd@exmple.com",
        "first_name": "Jan R.",
        "last_name": "Dressler",
        "affiliation": "Asia-Africa-Institute, University of Hamburg, Germany",
        "proposals": [
            {
                "title": "Shared Legacies – The Legendary History of Angkor and Its Nineteenth Century Genesis",
                "type": "paper",
                "abstract": "The magnificent ruins of Angkor Wat and Angkor Thom bear witness to the ancient \
civilization of Cambodia, which flourished at the banks of the Mekong river and the Tonlé Sap \
between the 9th and 13th century. It was, however, not until King Norodom I of Cambodia had \
mandated the reconstitution of the chronicles of his kingdom in 1869 AD, that Khmer language \
versions of the royal chronicles were compiled which, seemingly without any prior model, attempted \
to incorporate into the history of the Cambodian state the forgotten past of the Angkorean kingdom. \
Among the specimen of traditional Cambodian historiography preserved in the Siamese language alone, \
a text commonly referred to as “Rueang Phra Chao Prathum Suriwong” demonstrably predates this \
development. Based on Southeast Asian folklore and literature, and in a fashion less elaborate and \
refined as its successors, this short chronicle relates the legendary history of the Angkorean \
polity from the construction of the famous temple complexes throughout the ages until the 14th \
century. Drawing on a survey of Siamese archival and historiographic sources, I argue in this \
presentation that in the early 1860s the Cambodian royal court initially produced this first \
draft of a legendary history of the Angkorean kingdom on the explicit request of King Rama IV of \
Siam, and subsequently adapted the idea to suit its own political purposes.",
            }
        ],
    },
]


def get_engine() -> Engine:
    """Initialize engine.

    :rtype: sqlalchemy.engine.Engine
    """
    if "engine" not in g:
        g.engine = create_engine(
            current_app.config["SQLALCHEMY_DATABASE_URI"], echo=False, future=True
        )
    return g.engine


def close_engine(e: Any = None) -> None:
    """Close the database connection."""
    engine = g.pop("engine", None)
    if engine is not None:
        engine.dispose()


def init_db() -> None:
    """Create table."""
    engine = get_engine()

    from mngt.models import Base

    Base.metadata.create_all(engine)


def row_is_all_none_or_empty(row: Tuple) -> bool:
    """True if the entire row is None or an empty string."""
    return all(
        col.value is None or (isinstance(col.value, str) and len(col.value) == 0)
        for col in row
    )


def get_short_title(title: str) -> str:
    """Get the short version of a text."""
    titles = [title, title.split("\n")[0], title.split(".")[0]]
    lens = [len(title) for title in titles]
    short_titles = sorted(zip(lens, titles), key=lambda v: v[0])
    return short_titles[0][1]


def import_cots2021_proposals(file_path: Path, worksheet_name: str = None) -> None:
    """Import the COTS 2021's Excel Containing Proposals.

    :param file_path: The path to the Excel file.
    :param worksheet_name: The name of the worksheet to import.

    This assumes that the tables are already created.
    """
    from . import models

    click.echo("Opening the workbook ...")
    wb = load_workbook(file_path, data_only=True)

    if len(wb.worksheets) == 0:
        raise Exception("No worksheet found.")

    if worksheet_name is not None:
        if worksheet_name not in wb.worksheets:
            raise Exception(f"Worksheet {worksheet_name} not found.")
        ws = wb.get_sheet_by_name(worksheet_name)
    else:
        ws = wb.worksheets[0]

    # For each row, create participant object
    # if it does not already exist.
    # Then create the proposal object.

    participants = []
    participant_emails = []

    engine = get_engine()
    with Session(engine, future=True) as session:
        # NOTE: Hard coded value fot COTS 2021.
        cot2021_conf = models.Conference(
            name=_conferences[0]["name"],
            description=_conferences[0]["description"],
            slug="cots-2021",
            begin=_conferences[0]["begin"],
            end=_conferences[0]["end"],
            created=datetime.utcnow(),
            modified=datetime.utcnow(),
        )
        session.add(cot2021_conf)
        session.commit

        # min_row is 1-based index.
        for row in ws.iter_rows(min_row=2):
            # if row.participant.email not in participant_emails
            # create new object.
            if row_is_all_none_or_empty(row):
                continue

            if row[1].value is None:
                print([col.value for col in row])

            timestamp, email_address = row[0].value, row[1].value  # noqa: F841
            first_name, last_name = row[2].value, row[3].value  # noqa: F841
            affiliation, proposal_type = row[4].value, row[5].value  # noqa: F841

            if email_address in participant_emails:
                participant = None
                for p in participants:
                    if p.email == email_address:
                        participant = p
                        break
            else:
                participant = models.Participant(
                    email=email_address,
                    first_name=first_name,
                    last_name=last_name,
                    affiliation="" if affiliation is None else affiliation,
                    created=datetime.utcnow(),
                    modified=datetime.utcnow(),
                )
                participants.append(participant)
                participant_emails.append(email_address)
                session.add(participant)

            # create new proposal objets.
            if (
                proposal_type
                == "A proposal for an individual paper, film screening or other presentationtation"
            ):
                # Individual presentation.
                print(f"Process individual presentation proposal from {email_address}")

                abstract = row[6].value
                first_line = abstract.split("\n")[0]
                participant.proposals.append(
                    models.Proposal(
                        conference=cot2021_conf,
                        created=datetime.utcnow(),
                        modified=datetime.utcnow(),
                        title=first_line,
                        type=proposal_type,
                        abstract=abstract,
                    )
                )
            elif proposal_type == "A proposal for a paper panel":
                # Panel presentation.
                print(f"Process panel presentation proposal from {email_address}")

                panel_topic = row[7].value
                panelist_names = row[8].value  # noqa: F841
                panelist_emails = row[9].value  # noqa: F841
                panelist_abstracts = row[10].value

                participant.proposals.append(
                    models.Proposal(
                        conference=cot2021_conf,
                        created=datetime.utcnow(),
                        modified=datetime.utcnow(),
                        title=panel_topic,
                        type=proposal_type,
                        abstract=panelist_abstracts,
                    )
                )
            elif proposal_type == "A proposal for a routable":
                # Roundtable.
                print(f"Process roundtable presentation proposal from {email_address}")

                roundtable_names = row[11].value
                roundtable_emails = row[12].value  # noqa: F841

                short_title = get_short_title(roundtable_names)
                abstract = ""
                if short_title != roundtable_names:
                    abstract = roundtable_names

                participant.proposals.append(
                    models.Proposal(
                        conference=cot2021_conf,
                        created=datetime.utcnow(),
                        modified=datetime.utcnow(),
                        title=short_title,
                        type=proposal_type,
                        abstract=abstract,
                    )
                )
        session.commit()


@click.command("init-db")
@with_appcontext
def init_db_comamnd() -> None:
    """Create tables."""
    init_db()
    click.echo("Initialized the database.")


@click.command("seed-db")
@with_appcontext
def seed_db_command() -> None:
    """Seed the table with some data."""
    from . import models

    click.echo("Seeding the database ...")
    engine = get_engine()
    with Session(engine, future=True) as session:
        for item in _conferences:
            conf = models.Conference(
                name=item["name"],
                description=item["description"],
                begin=item["begin"],
                end=item["end"],
            )
            session.add(conf)

        proposal_objects = []
        for item in _participants:
            participant = models.Participant(
                email=item["email"],
                first_name=item["first_name"],
                last_name=item["last_name"],
                affiliation=item["affiliation"],
                created=datetime.utcnow(),
                modified=datetime.utcnow(),
            )
            session.add(participant)

            for proposal_item in item["proposals"]:
                participant.proposals.append(
                    models.Proposal(
                        created=datetime.utcnow(),
                        modified=datetime.utcnow(),
                        title=proposal_item["title"],
                        type=proposal_item["type"],
                        abstract=proposal_item["abstract"],
                    )
                )

        session.bulk_save_objects(proposal_objects)
        session.commit()


@click.command("import-cots2021")
@click.argument("filepath", type=click.Path(exists=True))
@with_appcontext
def import_cots2021_proposals_command(filepath: str) -> None:
    """
    A command to import Excel file from Google Form response.

    :param filepath: The path to Excel file.
    """
    if isinstance(filepath, str):
        filepath = Path(filepath)
    if not filepath.is_file():
        click.echo("Supplied path is not a file.")
        sys.exit(-1)

    import_cots2021_proposals(filepath)


def init_app(app: Flask) -> None:
    """Initialize application."""
    app.teardown_appcontext(close_engine)
    app.cli.add_command(init_db_comamnd)
    app.cli.add_command(seed_db_command)
    app.cli.add_command(import_cots2021_proposals_command)
